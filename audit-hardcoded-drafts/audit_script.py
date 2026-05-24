"""
ADVANCE hardcoded draft audit.

Runs on the server. Audits every draft in Drafts/hardcoded/{part,part_PC}/*
against:
  - participant intervention (is the draft in the right folder?)
  - sheet-row consistency (is the participant's active row the expected one?)
  - stamp columns (is the Excel stamp aligned with the draft?)
  - body content (does the template match the intervention?)
  - sent items (is this draft a duplicate of something already sent?)
  - dead-row / signed-out / superseded status

Prints a categorised report. Does not delete anything.

Usage (on server):
    cd email_draft_automation
    .venv/bin/python3 /tmp/audit_script.py

Or to scope to specific template families:
    .venv/bin/python3 /tmp/audit_script.py --only prog_info,prog_start
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

ROOT = Path(os.path.expanduser("~/email_draft_automation"))
sys.path.insert(0, str(ROOT / "src"))

from dotenv import load_dotenv  # noqa: E402
load_dotenv(ROOT / ".env")

import openpyxl  # noqa: E402
from graph_mail import build_graph_service_from_env  # noqa: E402

WORKBOOK = ROOT / "Qualtrics_ADVANCE_recruitment_live.xlsx"
TEMPLATES_PATH = ROOT / "context" / "templates" / "participant_email_templates.json"

# Folder name → expected template_key (used to detect mismatches)
INTERVENTION_FOLDERS = {
    "prog_info_standard":            {"POD", "SH+", "COG"},
    "prog_info_combined":            {"SH+/COG", "SH+ & COG", "SH+COG"},
    "prog_info_wcg":                 {"WCG"},
    "prog_start-in-1-week_pod":      {"POD"},
    "prog_start-in-1-week_sh":       {"SH+"},
    "prog_start-in-1-week_cog":      {"COG"},
    "prog_start-in-1-week_combined": {"SH+/COG", "SH+ & COG", "SH+COG"},
}

TREATMENT_FOLDERS = {
    "prog_info_standard", "prog_info_combined",
    "prog_start-in-1-week_pod", "prog_start-in-1-week_sh",
    "prog_start-in-1-week_cog", "prog_start-in-1-week_combined",
}

# Folder → Excel stamp column
STAMP_COLUMNS = {
    "registration_confirmation": "registration_confirmation_draft_at",
    "consent_pre_invite":        "consent_pre_invite_draft_at",
    "consent_pre_reminder_1":    "consent_pre_reminder_1_draft_at",
    "consent_pre_reminder_2":    "consent_pre_reminder_2_draft_at",
    "prog_info_standard":        "prog_info_draft_at",
    "prog_info_combined":        "prog_info_draft_at",
    "prog_info_wcg":             "prog_info_draft_at",
    "prog_start-in-1-week_pod":  "prog_reminder_draft_at",
    "prog_start-in-1-week_sh":   "prog_reminder_draft_at",
    "prog_start-in-1-week_cog":  "prog_reminder_draft_at",
    "prog_start-in-1-week_combined": "prog_reminder_draft_at",
    "signout_confirmation":      "signout_confirmation_draft_at",
}

# Text markers for body content identification
BODY_MARKERS = {
    "wcg": [r"liste d[’']attente", r"Warteliste", r"aléatoirement", r"zugeteilt"],
    "standard_biweekly": [r"toutes les deux semaines", r"alle zwei Wochen"],
    "combined_weekly": [r"une fois par semaine", r"einmal pro Woche"],
    "pod_platform_error": [r"plateforme en ligne", r"Online[- ]?Plattform"],
    "unresolved_placeholder": [r"\{[a-z_]+\}"],
}


def cv(v) -> str:
    return "" if v is None else str(v).strip()


def parse_d(s: str):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def norm_subject(s: str) -> str:
    return re.sub(r"^(Re:\s*|RE:\s*|Fwd?:\s*|FW:\s*)+", "", s, flags=re.IGNORECASE).strip().lower()


def load_participants() -> dict:
    """Return email → list of rows (one per sheet occurrence). Keeps active + superseded + inactive + interested-person.

    Glossary (docs/glossary.md): after partitioning, dead rows (signed_out,
    bounced, superseded) move from part/part_PC to part_inactive. part_int
    lives in a separate bounded context — its signed_out flag has different
    semantics. This loader picks up all four sheets so the audit sees the
    full lifecycle across contexts. The part_int email column is named
    `part_int_email`, not `part_email`.
    """
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    out: dict = defaultdict(list)
    sheets = ("part", "part_PC", "part_inactive", "part_int")
    email_col_by_sheet = {
        "part": "part_email", "part_PC": "part_email",
        "part_inactive": "part_email", "part_int": "part_int_email",
    }
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cv(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        email_col = email_col_by_sheet.get(sheet_name, "part_email")
        ei = headers.index(email_col) if email_col in headers else None
        if ei is None:
            continue
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            email = cv(row[ei].value).lower()
            if not email:
                continue
            vals = {headers[i]: row[i].value for i in range(len(headers)) if i < len(row)}
            out[email].append({"sheet": sheet_name, "row_num": row_num, "vals": vals, "headers": headers})
    wb.close()
    return out


# Folders that are driven by the part_int sheet rather than part/part_PC.
# For these, an active part_int row is sufficient — the recipient doesn't
# need an active part or part_PC row.
PART_INT_FOLDERS = {
    "part_int_reminder_1", "part_int_reminder_2",
    "part_int_outreach", "part_int_invite",
}


def active_row_for(email: str, participants: dict) -> dict | None:
    """Return the single active (non-superseded, non-dead) row for an email.

    Preference: active part row > active part_PC row. part_inactive rows are
    never considered active (glossary: is_dead_row is true by construction
    for every row on that sheet).
    """
    rows = [r for r in participants.get(email, []) if r["sheet"] != "part_inactive"]
    # Filter out superseded + postmaster-bounced
    candidates = [r for r in rows if not cv(r["vals"].get("part_superseded_by"))
                  and cv(r["vals"].get("email_postmaster_failure") or "").upper() != "TRUE"
                  and cv(r["vals"].get("signed_out") or "").upper() != "TRUE"]
    # Prefer enrolled rows
    part_enrolled = [r for r in candidates if r["sheet"] == "part"
                     and cv(r["vals"].get("part_enrolled") or "").upper() == "TRUE"]
    if part_enrolled:
        return part_enrolled[0]
    pc_enrolled = [r for r in candidates if r["sheet"] == "part_PC"
                   and cv(r["vals"].get("part_PC_enrolled") or "").upper() == "TRUE"]
    if pc_enrolled:
        return pc_enrolled[0]
    return candidates[0] if candidates else None


def get_folder_id(mail, mailbox: str, path: list[str]) -> str | None:
    resp = mail._request("GET", f"/users/{mailbox}/mailFolders", params={"$top": "50"})
    folders = {f["displayName"]: f["id"] for f in resp.json().get("value", [])}
    fid = folders.get("Drafts")
    for p in path:
        if not fid:
            return None
        resp = mail._request("GET", f"/users/{mailbox}/mailFolders/{fid}/childFolders",
                             params={"$top": "100"})
        sub = {f["displayName"]: f["id"] for f in resp.json().get("value", [])}
        fid = sub.get(p)
    return fid


def list_child_folders(mail, mailbox: str, parent_fid: str) -> list[dict]:
    resp = mail._request("GET", f"/users/{mailbox}/mailFolders/{parent_fid}/childFolders",
                         params={"$top": "100"})
    return resp.json().get("value", [])


def get_messages(mail, mailbox: str, fid: str) -> list[dict]:
    items = []
    url = f"/users/{mailbox}/mailFolders/{fid}/messages"
    params = {"$top": "100", "$select": "toRecipients,subject,id,body"}
    while url:
        resp = mail._request("GET", url, params=params)
        data = resp.json()
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
        params = None
    return items


def crawl_hardcoded(mail, mailbox: str, only_families: set[str] | None = None) -> list[dict]:
    """Return list of {email, folder (role/name), id, subject, body}."""
    drafts: list[dict] = []
    hardcoded_fid = get_folder_id(mail, mailbox, ["hardcoded"])
    if not hardcoded_fid:
        print("!! hardcoded/ folder not found under Drafts/")
        return drafts

    for role_f in list_child_folders(mail, mailbox, hardcoded_fid):
        role = role_f["displayName"]
        if role == "facilities":
            continue  # not in this audit's scope
        for tmpl_f in list_child_folders(mail, mailbox, role_f["id"]):
            folder_name = tmpl_f["displayName"]
            if only_families:
                if not any(folder_name.startswith(p) for p in only_families):
                    continue
            for m in get_messages(mail, mailbox, tmpl_f["id"]):
                for recip in m.get("toRecipients", []):
                    addr = cv(recip.get("emailAddress", {}).get("address", "")).lower()
                    drafts.append({
                        "email": addr,
                        "role": role,  # "part" or "part_PC"
                        "folder": folder_name,
                        "id": m["id"],
                        "subject": cv(m.get("subject")),
                        "body": m.get("body", {}).get("content", "") or "",
                    })
    return drafts


def fetch_sent(mail, mailbox: str, email: str) -> list[tuple[str, str]]:
    try:
        resp = mail._request("GET", f"/users/{mailbox}/mailFolders/SentItems/messages",
                             params={"$search": f'"to:{email}"', "$top": "20",
                                     "$select": "subject,sentDateTime"})
        return [(cv(m.get("subject")), cv(m.get("sentDateTime"))) for m in resp.json().get("value", [])]
    except Exception:
        return []


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--only", default="",
                        help="Comma-separated folder-name prefixes (e.g. prog_info,prog_start)")
    args = parser.parse_args()
    only_families = {p.strip() for p in args.only.split(",") if p.strip()} or None

    mail = build_graph_service_from_env()
    mailbox = os.environ["MS_MAILBOX"]

    print("Loading participants from workbook...")
    participants = load_participants()
    print(f"  {len(participants)} unique emails, "
          f"{sum(len(v) for v in participants.values())} total rows")

    print("Crawling hardcoded draft folders...")
    drafts = crawl_hardcoded(mail, mailbox, only_families)
    print(f"  {len(drafts)} drafts found")

    issues: dict = defaultdict(list)
    today = date.today()

    # ── Cache sent items per email (only those with drafts) ──────────────
    print("Fetching sent items for each recipient...")
    sent_cache: dict = {}
    for email in {d["email"] for d in drafts}:
        sent_cache[email] = fetch_sent(mail, mailbox, email)

    # ── Per-draft checks ────────────────────────────────────────────────
    for d in drafts:
        email = d["email"]
        folder = d["folder"]
        role = d["role"]
        subject = d["subject"]
        body = d["body"]
        body_text = re.sub(r"<[^>]+>", " ", body)
        body_text = re.sub(r"\s+", " ", body_text).strip()

        # part_int folders are driven by the part_int sheet (bounded context),
        # not part / part_PC. A draft here is legit if the participant has
        # an active (non-signed-out, non-bounced) part_int row, regardless
        # of their part / part_PC status.
        if folder in PART_INT_FOLDERS:
            part_int_rows = [r for r in participants.get(email, [])
                             if r["sheet"] == "part_int"
                             and cv(r["vals"].get("signed_out") or "").upper() != "TRUE"
                             and cv(r["vals"].get("email_postmaster_failure") or "").upper() != "TRUE"]
            if part_int_rows:
                continue  # legit part_int draft — no further checks apply
            # No active part_int row → flag as inactive-only if they're on part_int_inactive
            # (or whatever partition sheet exists) or unknown otherwise.
            issues["DRAFT_FOR_INACTIVE_ONLY"].append({
                "email": email, "folder": folder,
                "note": "no active part_int row",
            })
            continue

        row = active_row_for(email, participants)
        if row is None:
            if email.endswith("@unige.ch") or email.endswith("@phbern.ch"):
                continue  # likely facility, but facilities are excluded above
            # Distinguish: recipient on part_inactive only vs truly unknown
            has_inactive = any(r["sheet"] == "part_inactive"
                               for r in participants.get(email, []))
            if has_inactive:
                reasons = sorted({cv(r["vals"].get("part_inactive_reason"))
                                  for r in participants.get(email, [])
                                  if r["sheet"] == "part_inactive"})
                issues["DRAFT_FOR_INACTIVE_ONLY"].append({
                    "email": email, "folder": folder,
                    "inactive_reasons": [r for r in reasons if r],
                })
            else:
                issues["UNKNOWN_EMAIL"].append({"email": email, "folder": folder})
            continue

        vals = row["vals"]
        sheet = row["sheet"]
        intervention = cv(vals.get("part_intervention"))
        effective_interv = intervention.upper() if sheet == "part" else "WCG"

        # Check if active row would satisfy base creation conditions for this folder
        enrolled = cv(vals.get("part_PC_enrolled" if sheet == "part_PC" else "part_enrolled")).upper() == "TRUE"
        if folder != "signout_confirmation" and not enrolled:
            issues["DRAFT_FOR_NOT_ENROLLED"].append({
                "email": email, "folder": folder, "sheet": sheet,
                "reason": "active row is part_enrolled=FALSE — code wouldn't recreate",
            })

        # 1. WRONG_INTERVENTION
        if folder in INTERVENTION_FOLDERS:
            allowed = {a.upper() for a in INTERVENTION_FOLDERS[folder]}
            if effective_interv and effective_interv not in allowed:
                issues["WRONG_INTERVENTION"].append({
                    "email": email, "folder": folder,
                    "actual": effective_interv, "expected_any_of": sorted(allowed),
                })

        # 2. WRONG_SHEET (role vs active sheet)
        if role == "part" and sheet == "part_PC":
            issues["WRONG_SHEET"].append({"email": email, "folder": folder,
                                          "draft_role": role, "active_sheet": sheet})
        elif role == "part_PC" and sheet == "part":
            issues["WRONG_SHEET"].append({"email": email, "folder": folder,
                                          "draft_role": role, "active_sheet": sheet})

        # 3. WCG_GOT_TREATMENT_DRAFT
        if sheet == "part_PC" and folder in TREATMENT_FOLDERS:
            issues["WCG_GOT_TREATMENT_DRAFT"].append({"email": email, "folder": folder})

        # 4. TREATMENT_GOT_WCG_DRAFT
        if sheet == "part" and folder == "prog_info_wcg":
            issues["TREATMENT_GOT_WCG_DRAFT"].append({
                "email": email, "folder": folder, "intervention": intervention,
            })

        # 5. DRAFT_WITHOUT_STAMP
        stamp_col = STAMP_COLUMNS.get(folder)
        if stamp_col:
            stamp = cv(vals.get(stamp_col))
            if not stamp:
                issues["DRAFT_WITHOUT_STAMP"].append({
                    "email": email, "folder": folder, "stamp_col": stamp_col,
                })

        # 8. STAMP_PREDATES_CONFIRMATION (only for consent_pre_invite)
        if folder == "consent_pre_invite":
            inv_stamp = parse_d(cv(vals.get("consent_pre_invite_draft_at")))
            conf = parse_d(cv(vals.get("registration_confirmation_draft_at")))
            if inv_stamp and conf and inv_stamp < conf:
                issues["STAMP_PREDATES_CONFIRMATION"].append({
                    "email": email, "inv_stamp": str(inv_stamp), "conf": str(conf),
                })

        # 9-11. Content checks
        body_lower = body_text.lower()
        for pat in BODY_MARKERS["unresolved_placeholder"]:
            m = re.search(pat, body_text)
            if m:
                issues["MISSING_PLACEHOLDER"].append({
                    "email": email, "folder": folder, "placeholder": m.group(0),
                })
                break

        if folder.startswith("prog_info"):
            has_wcg = any(re.search(p, body_text, re.IGNORECASE) for p in BODY_MARKERS["wcg"])
            has_biweekly = any(re.search(p, body_text, re.IGNORECASE) for p in BODY_MARKERS["standard_biweekly"])
            has_weekly = any(re.search(p, body_text, re.IGNORECASE) for p in BODY_MARKERS["combined_weekly"])
            has_platform = any(re.search(p, body_text, re.IGNORECASE) for p in BODY_MARKERS["pod_platform_error"])

            if folder == "prog_info_wcg" and not has_wcg:
                issues["TEMPLATE_CONTENT_MISMATCH"].append({
                    "email": email, "folder": folder, "reason": "wcg body without wait-list marker",
                })
            if folder == "prog_info_standard" and not has_biweekly:
                issues["TEMPLATE_CONTENT_MISMATCH"].append({
                    "email": email, "folder": folder, "reason": "standard body missing 'toutes les deux semaines'",
                })
            if folder == "prog_info_combined" and not has_weekly:
                issues["TEMPLATE_CONTENT_MISMATCH"].append({
                    "email": email, "folder": folder, "reason": "combined body missing 'une fois par semaine'",
                })
            if folder == "prog_info_standard" and effective_interv == "POD" and has_platform:
                issues["POD_WITH_PLATFORM_TEXT"].append({"email": email, "folder": folder})

        # 12. ALREADY_SENT
        sent_msgs = sent_cache.get(email, [])
        draft_subj_norm = norm_subject(subject)
        for s_subj, s_date in sent_msgs:
            if norm_subject(s_subj) == draft_subj_norm:
                issues["ALREADY_SENT"].append({
                    "email": email, "folder": folder, "sent_date": s_date, "subject": subject[:60],
                })
                break

        # 15. DRAFT_FOR_SIGNED_OUT
        signed = cv(vals.get("signed_out") or "").upper() == "TRUE"
        if signed and folder != "signout_confirmation":
            issues["DRAFT_FOR_SIGNED_OUT"].append({"email": email, "folder": folder})

        # 16. DRAFT_FOR_DEAD_ROW — can't happen via active_row_for since it filters,
        # but we separately check the original rows for the email
        for r in participants.get(email, []):
            if cv(r["vals"].get("part_superseded_by")) and r["sheet"] == role:
                issues["DRAFT_FOR_DEAD_ROW"].append({
                    "email": email, "folder": folder, "role": role,
                    "reason": "matching sheet row is superseded",
                })
                break

    # ── Excel-side checks (stamps without drafts / out-of-order) ────────
    print("Checking stamp chronology...")
    for email, rows in participants.items():
        for r in rows:
            if cv(r["vals"].get("part_superseded_by")):
                continue
            rem1 = parse_d(cv(r["vals"].get("consent_pre_reminder_1_draft_at")))
            rem2 = parse_d(cv(r["vals"].get("consent_pre_reminder_2_draft_at")))
            if rem1 and rem2:
                gap = (rem2 - rem1).days
                if gap < 0:
                    issues["STAMP_OUT_OF_ORDER"].append({
                        "email": email, "rem1": str(rem1), "rem2": str(rem2),
                    })
                elif 0 <= gap < 5:
                    issues["REMINDER_GAP_TOO_SHORT"].append({
                        "email": email, "gap_days": gap, "rem1": str(rem1), "rem2": str(rem2),
                    })

            # Stamp predates conf (already handled per-draft, but also catch standalone stamps)
            conf = parse_d(cv(r["vals"].get("registration_confirmation_draft_at")))
            inv_stamp = parse_d(cv(r["vals"].get("consent_pre_invite_draft_at")))
            if conf and inv_stamp and inv_stamp < conf:
                issues["STAMP_PREDATES_CONFIRMATION"].append({
                    "email": email, "inv_stamp": str(inv_stamp), "conf": str(conf),
                })

    # ── Dual active rows ────────────────────────────────────────────────
    print("Checking dual-row integrity...")
    for email, rows in participants.items():
        # Exclude part_inactive by construction (every row there is dead)
        active = [r for r in rows
                  if r["sheet"] != "part_inactive"
                  and not cv(r["vals"].get("part_superseded_by"))
                  and cv(r["vals"].get("signed_out") or "").upper() != "TRUE"]
        sheets = {r["sheet"] for r in active}
        if len(sheets) > 1:
            issues["DUAL_ACTIVE_ROWS"].append({
                "email": email, "sheets": sorted(sheets),
                "rows": [(r["sheet"], r["row_num"]) for r in active],
            })

    # ── Missing-draft check (stamp set, no draft, no sent) ───────────────
    # Limited to stamps we care about for the invite phase
    print("Checking stamps without drafts...")
    draft_emails_by_folder = defaultdict(set)
    for d in drafts:
        draft_emails_by_folder[d["folder"]].add(d["email"])

    for email, rows in participants.items():
        for r in rows:
            if cv(r["vals"].get("part_superseded_by")):
                continue
            for folder, stamp_col in STAMP_COLUMNS.items():
                stamp = cv(r["vals"].get(stamp_col))
                if not stamp:
                    continue
                if email in draft_emails_by_folder.get(folder, set()):
                    continue
                # Check sent items
                sent_msgs = sent_cache.get(email)
                if sent_msgs is None:
                    sent_msgs = fetch_sent(mail, mailbox, email)
                    sent_cache[email] = sent_msgs
                # We don't know the exact subject to match; just check for any sent
                # whose subject "looks like" this template. Heuristic: subject contains
                # a keyword. Skip this heuristic to avoid false positives — only flag
                # when absolutely zero sent items to this recipient exist.
                if not sent_msgs:
                    issues["MISSING_DRAFT"].append({
                        "email": email, "stamp_col": stamp_col, "folder": folder,
                        "stamp_value": stamp,
                    })

    # ── Print report ────────────────────────────────────────────────────
    print()
    print("=" * 72)
    print("AUDIT REPORT")
    print("=" * 72)
    print(f"Drafts scanned: {len(drafts)}")
    print(f"Unique recipients: {len({d['email'] for d in drafts})}")
    print()

    order = [
        "WRONG_INTERVENTION", "WRONG_SHEET",
        "WCG_GOT_TREATMENT_DRAFT", "TREATMENT_GOT_WCG_DRAFT",
        "DRAFT_FOR_INACTIVE_ONLY", "DRAFT_FOR_NOT_ENROLLED",
        "DRAFT_WITHOUT_STAMP", "MISSING_DRAFT",
        "STAMP_OUT_OF_ORDER", "STAMP_PREDATES_CONFIRMATION",
        "MISSING_PLACEHOLDER", "POD_WITH_PLATFORM_TEXT", "TEMPLATE_CONTENT_MISMATCH",
        "ALREADY_SENT",
        "DUAL_ACTIVE_ROWS", "DRAFT_FOR_SIGNED_OUT", "DRAFT_FOR_DEAD_ROW",
        "REMINDER_GAP_TOO_SHORT",
        "UNKNOWN_EMAIL",
    ]
    total = 0
    for k in order:
        n = len(issues.get(k, []))
        total += n
        marker = "" if n == 0 else " !!"
        print(f"  {k:<32} {n}{marker}")
    print(f"  {'TOTAL':<32} {total}")
    print()

    for k in order:
        entries = issues.get(k, [])
        if not entries:
            continue
        print(f"--- {k} ({len(entries)}) ---")
        for e in entries[:50]:
            print(f"  {json.dumps(e, sort_keys=True)}")
        if len(entries) > 50:
            print(f"  ... and {len(entries) - 50} more")
        print()

    print("Done.")


if __name__ == "__main__":
    main()
