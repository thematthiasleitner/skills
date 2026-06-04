#!/usr/bin/env python3
"""Census date/time/number/text format shapes across every sheet of an ADVANCE workbook.

Read-only. The before/after verification engine for the date/time standardisation
(ADR-pending). Surfaces which columns are format-MIXED and splits the two families:

  Family A — course/event dates+times humans read & emails render
             (begin/time, part_prog_*, course_*, session_date, admin_rec_date, trai_*)
             → SHOULD be one canonical format (dd/mm/yyyy + HH:MM).
  Family B — machine audit stamps (*_at, *_sent_at, *_logged_at, recorded_date, *_fini_date)
             → ISO on purpose; drive dedup/sorting; LEAVE ALONE.

CLI:
  _format_census.py                                  # live workbook, date/time cols, MIXED-focus
  _format_census.py --workbook ..._preview.xlsx
  _format_census.py --all-shapes                     # also show non-mixed columns
  _format_census.py --family A                        # only Family-A columns
"""
from __future__ import annotations

import argparse
import re
from collections import Counter

from openpyxl import load_workbook

DEFAULT_WORKBOOK = "Qualtrics_ADVANCE_recruitment_live.xlsx"

KEY = re.compile(r"date|begin|datum|time|hour|heure|uhr|zeit|_at$|day|session|recorded|fini", re.I)
# Family B = machine timestamps (ISO, leave alone)
FAMILY_B = re.compile(r"_at$|_sent_at$|_logged_at$|_draft_at$|recorded_date$|_fini_date$|received_at$|inactive_at$|reclassified_at$", re.I)
# duration-like columns that are NUMBERS, not clock times — never "time-normalise"
DURATION = re.compile(r"hours|_hours|remain|weeks|_cap$|_N_t$", re.I)


def dshape(s: str) -> str | None:
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", s): return "D:d/m/Y"
    if re.fullmatch(r"\d{1,2}\.\d{1,2}\.\d{4}", s): return "D:d.m.Y"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", s): return "D:ISO"
    if re.fullmatch(r"\d{1,2}-\d{1,2}-\d{4}", s): return "D:d-m-Y"
    return None


def tshape(s: str) -> str | None:
    if re.fullmatch(r"\d{1,2}:\d{2}", s): return "T:H:MM"
    if re.fullmatch(r"\d{1,2}\.\d{2}", s): return "T:H.MM"
    if re.fullmatch(r"\d{1,2}h\d{0,2}", s): return "T:Hh"
    if re.fullmatch(r"\d{1,2}", s): return "T:bareH"
    return None


def family(col: str) -> str:
    return "B" if FAMILY_B.search(col) else "A"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    ap.add_argument("--all-shapes", action="store_true", help="show non-mixed columns too")
    ap.add_argument("--family", choices=["A", "B"], help="restrict to one family")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    print(f"=== FORMAT CENSUS: {args.workbook} ===")
    print("Family A = course/event (normalise to dd/mm/yyyy + HH:MM) | Family B = ISO stamps (leave)\n")
    mixed_a = 0
    for sh in wb.sheetnames:
        ws = wb[sh]
        try:
            hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
        except StopIteration:
            continue
        cand = [(i, h) for i, h in enumerate(hdr) if h and KEY.search(str(h)) and not DURATION.search(str(h))]
        if not cand:
            continue
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r and not all(v in (None, "") for v in r)]
        lines = []
        for i, h in cand:
            fam = family(str(h))
            if args.family and fam != args.family:
                continue
            shapes = Counter()
            for r in rows:
                v = r[i] if i < len(r) else None
                if v in (None, ""):
                    continue
                s = dshape(str(v)) or tshape(str(v))
                if s:
                    shapes[s] += 1
            if not shapes:
                continue
            is_mixed = len({k.split(":")[0] + k for k in shapes}) > 1 and len(shapes) > 1
            if fam == "A" and len(shapes) > 1:
                mixed_a += 1
            if is_mixed or args.all_shapes:
                tag = "  <== MIXED" if len(shapes) > 1 else ""
                lines.append(f"  [{fam}] {h:34} {dict(shapes)}{tag}")
        if lines:
            print(f"### {sh} ({len(rows)} rows)")
            print("\n".join(lines))
            print()
    wb.close()
    print(f"Family-A columns that are format-MIXED: {mixed_a}")
    if mixed_a == 0 and not args.all_shapes:
        print("(none — Family A is uniform. Run with --all-shapes to see the canonical they agree on.)")


if __name__ == "__main__":
    main()
