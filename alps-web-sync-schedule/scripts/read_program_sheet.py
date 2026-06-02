#!/usr/bin/env python3
"""Dump a planning-sheet tab laid out as per-day 3-column blocks (time | who | topic).

Each day occupies three adjacent columns under a header cell that names the day
(a weekday, a date, or "Day N"). A blank per-day time cell falls back to the
"shared" leftmost time column (the column whose cells look most like "HH:MM - HH:MM").

Usage:
  read_program_sheet.py FILE.xlsx [--sheet NAME] [--day SUBSTRING]

  --sheet  worksheet name (default: the first sheet that has day headers, else the first sheet)
  --day    only print days whose header contains this (case-insensitive), e.g. "Thursday"

Install openpyxl if missing:
  python3 -m pip install --user --break-system-packages openpyxl
"""
import argparse
import re
import sys
import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl missing — run: python3 -m pip install --user --break-system-packages openpyxl")

DAY_RE = re.compile(
    r"(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday|Day\s*\d)", re.I
)
TIME_RE = re.compile(r"\d{1,2}[:.]\d{2}\s*[-–]\s*\d{1,2}[:.]\d{2}")


def cell(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.time, datetime.datetime, datetime.date)):
        return v.isoformat()
    return str(v).strip()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--sheet")
    ap.add_argument("--day")
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.file, read_only=True, data_only=True)

    def header_cols(ws):
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append([cell(c) for c in r])
            if i >= 4:
                break
        for ri, row in enumerate(rows):
            hits = [(ci, txt) for ci, txt in enumerate(row) if DAY_RE.search(txt)]
            if hits:
                return ri, hits
        return None, []

    # pick sheet
    ws = None
    if a.sheet:
        ws = wb[a.sheet]
    else:
        for cand in wb.worksheets:
            if header_cols(cand)[1]:
                ws = cand
                break
        ws = ws or wb.worksheets[0]

    rows = [[cell(c) for c in r] for r in ws.iter_rows(values_only=True)]
    hdr_ri, hits = header_cols(ws)
    if not hits:
        sys.exit(f"No day headers found in sheet '{ws.title}'. Sheets: {wb.sheetnames}")

    # shared time column = column with the most time-pattern cells
    ncols = max(len(r) for r in rows)
    best_col, best_n = 0, -1
    for ci in range(ncols):
        n = sum(1 for r in rows if ci < len(r) and TIME_RE.search(r[ci]))
        if n > best_n:
            best_col, best_n = ci, n

    print(f"# sheet: {ws.title}   shared-time col: {best_col}")
    for ci, label in hits:
        if a.day and a.day.lower() not in label.lower():
            continue
        print(f"\n## {label}  (cols {ci}..{ci+2})")
        for r in rows[hdr_ri + 1:]:
            t = r[ci] if ci < len(r) else ""
            who = r[ci + 1] if ci + 1 < len(r) else ""
            topic = r[ci + 2] if ci + 2 < len(r) else ""
            if not t and (who or topic):  # fall back to shared time column
                t = r[best_col] if best_col < len(r) else ""
            if t or who or topic:
                print(f"  {t:<16} | {who:<22} | {topic}")


if __name__ == "__main__":
    main()
