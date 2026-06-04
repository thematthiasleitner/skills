#!/usr/bin/env python3
"""Trace an ADVANCE column end-to-end: survey question -> CSV tag -> export code -> workbook.

The investigative loop that the f_ty_2_con audit and the dead-column audit both
hand-rolled. Read-only. Run from inside email_draft_automation.

Answers, for a given column/tag, the four questions a column audit needs:
  1. SURVEY  — is the source question still alive? (QID, type, validation, block, orphaned?)
  2. CSV     — which CSV columns carry it, and what values come in?
  3. CODE    — where does export code reference it? (file:line hits)
  4. WORKBOOK — which sheet(s) hold it, and what's the value distribution / format?

Artifacts are REUSED if already present (newest dated file in qualtrics/), else
pulled fresh — so a trace is fast when you've just pulled, fresh when you haven't.

CLI:
  _column_trace.py f_ty_room_tn_begin\\(n\\)
  _column_trace.py part_zoom --pull          # force fresh survey + CSV
  _column_trace.py admin_rec_date --workbook Qualtrics_ADVANCE_recruitment_live.xlsx
"""
from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import sys
from collections import Counter
from pathlib import Path

QUALTRICS_DIR = Path("qualtrics")
DEFAULT_WORKBOOK = "Qualtrics_ADVANCE_recruitment_live.xlsx"
CODE_GLOBS = ["qualtrics/*.py", "src/*.py"]

sys.path.insert(0, str(Path.home() / ".claude/skills/qualtrics-survey-pull"))


def _newest(pattern: str) -> str | None:
    hits = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    return hits[0] if hits else None


def _core(name: str) -> str:
    """Normalise a column/tag to a comparable core: drop leading 'N_'/'<digit>_',
    trailing '(n)'/'(N)', and any '_N_'/'_t<digit>_' loop infixes."""
    s = name.strip()
    s = re.sub(r"^\d+_", "", s)
    s = re.sub(r"^N_", "", s)
    s = re.sub(r"\((?:n|N)\)$", "", s)
    s = s.replace("_N_", "_").replace("tn_", "t_")
    s = re.sub(r"_t\d+_", "_t_", s)
    return s.lower().strip("_")


# ── 1. SURVEY ───────────────────────────────────────────────────────────────

def survey_section(column: str, pull: bool) -> str:
    import json
    from _survey_pull import question_report, question_by_tag, pull_survey_definition  # type: ignore
    saved = None if pull else _newest(str(QUALTRICS_DIR / "ADVANCE_recruitment_definition_*.json"))
    if saved:
        result = json.load(open(saved)).get("result", {})
        src = f"(from {saved})"
    else:
        result = pull_survey_definition()
        src = "(pulled fresh)"
    out = [f"## 1. SURVEY {src}"]
    rep = question_report(result, column)
    if not rep.get("found"):
        # try core-match across all tags
        cands = [q.get("DataExportTag") for q in result.get("Questions", {}).values()
                 if _core(q.get("DataExportTag", "")) == _core(column)]
        if cands:
            out.append(f"  no exact tag '{column}'; core-matches: {sorted(set(cands))}")
            rep = question_report(result, cands[0])
        else:
            out.append(f"  no survey question matches '{column}' (likely a DERIVED/internal workbook column, not a survey tag)")
            return "\n".join(out)
    out.append(f"  {rep['qid']}  type={rep['type']}  block={rep['block']}  in_flow={rep['in_flow']}")
    if rep.get("validation"):
        out.append(f"  validation={rep['validation']}")
    out.append(f"  text: {rep.get('question_text','')}")
    if "<NO BLOCK" in str(rep.get("block")):
        out.append("  ⚠️ ORPHANED — in no block, collects nothing from new respondents (DEAD source).")
    return "\n".join(out)


# ── 2. CSV ──────────────────────────────────────────────────────────────────

def _read_csv(path: str) -> tuple[list[str], list[list[str]]]:
    rows = list(csv.reader(open(path, encoding="utf-8-sig")))
    return rows[0], rows[3:]  # header row0; skip the 2 metadata header rows


def csv_section(column: str, pull: bool) -> str:
    out = ["## 2. CSV"]
    if pull:
        sys.path.insert(0, str(Path.home() / ".claude/skills/qualtrics-csv-pull"))
        try:
            from _csv_pull import pull_live_csv  # type: ignore
            rows = pull_live_csv()
            hdr = list(rows[0].keys()) if rows else []
            data = [[r.get(h, "") for h in hdr] for r in rows]
            src = "(pulled fresh)"
        except Exception as e:  # noqa: BLE001
            return f"## 2. CSV\n  could not pull fresh: {e}"
    else:
        path = _newest(str(QUALTRICS_DIR / "ADVANCE_recruitment_*_labels.csv")) or _newest(str(QUALTRICS_DIR / "ADVANCE_recruitment*.csv"))
        if not path:
            return "## 2. CSV\n  no saved labels CSV in qualtrics/ — re-run with --pull"
        hdr, data = _read_csv(path)
        src = f"(from {path})"
    out[0] = f"## 2. CSV {src}"
    core = _core(column)
    matched = [(i, h) for i, h in enumerate(hdr) if _core(h) == core or h == column]
    if not matched:
        out.append(f"  no CSV column matches core '{core}'")
        return "\n".join(out)
    for i, h in matched[:12]:
        vals = Counter((r[i].strip() if i < len(r) else "") for r in data)
        nonempty = sum(n for v, n in vals.items() if v)
        top = ", ".join(f"{v!r}×{n}" for v, n in vals.most_common(4) if v)
        out.append(f"  [{i}] {h}: {nonempty} non-empty | top: {top}")
    if len(matched) > 12:
        out.append(f"  …+{len(matched)-12} more matched columns")
    return "\n".join(out)


# ── 3. CODE ─────────────────────────────────────────────────────────────────

def code_section(column: str) -> str:
    out = ["## 3. CODE (export references)"]
    needle = column
    hits = []
    for g in CODE_GLOBS:
        for fp in glob.glob(g):
            try:
                for n, line in enumerate(open(fp, encoding="utf-8"), 1):
                    if needle in line:
                        hits.append(f"  {fp}:{n}: {line.strip()[:110]}")
            except Exception:  # noqa: BLE001
                continue
    if not hits:
        out.append(f"  no code references to '{column}' in {CODE_GLOBS}")
    else:
        out += hits[:18]
        if len(hits) > 18:
            out.append(f"  …+{len(hits)-18} more references")
    return "\n".join(out)


# ── 4. WORKBOOK ─────────────────────────────────────────────────────────────

def _shape(v) -> str:
    s = str(v)
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", s): return "d/m/Y"
    if re.fullmatch(r"\d{1,2}\.\d{1,2}\.\d{4}", s): return "d.m.Y"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", s): return "ISO"
    if re.fullmatch(r"\d{1,2}:\d{2}", s): return "H:MM"
    if re.fullmatch(r"\d{1,2}", s): return "int"
    return "text"


def workbook_section(column: str, wb_path: str) -> str:
    out = [f"## 4. WORKBOOK ({wb_path})"]
    if not Path(wb_path).exists():
        return out[0] + "\n  workbook not found"
    from openpyxl import load_workbook
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    found_any = False
    for sh in wb.sheetnames:
        ws = wb[sh]
        try:
            hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
        except StopIteration:
            continue
        idxs = [i for i, h in enumerate(hdr) if h == column or _core(str(h or "")) == _core(column)]
        for i in idxs:
            found_any = True
            vals = []
            shapes = Counter()
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r is None or all(v in (None, "") for v in r):
                    continue
                v = r[i] if i < len(r) else None
                if v in (None, ""):
                    continue
                vals.append(v)
                shapes[_shape(v)] += 1
            mix = "  <== MIXED" if len(shapes) > 1 else ""
            out.append(f"  {sh}.{hdr[i]}: {len(vals)} non-empty | shapes={dict(shapes)}{mix}")
    wb.close()
    if not found_any:
        out.append(f"  no sheet has a column matching '{column}'")
    return "\n".join(out)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("column", help="workbook column name OR CSV DataExportTag (e.g. 'f_ty_room_tn_begin(n)', 'part_zoom')")
    ap.add_argument("--pull", action="store_true", help="force fresh survey + CSV pulls instead of newest saved")
    ap.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    args = ap.parse_args()

    print(f"=== TRACE: {args.column}  (core='{_core(args.column)}') ===\n")
    print(survey_section(args.column, args.pull)); print()
    print(csv_section(args.column, args.pull)); print()
    print(code_section(args.column)); print()
    print(workbook_section(args.column, args.workbook))


if __name__ == "__main__":
    main()
